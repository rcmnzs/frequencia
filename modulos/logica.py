import sqlite3
import pandas as pd
import os
import re
from datetime import datetime
import locale

import fitz

# Importações para formatação do Excel
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Configura o locale para português
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
    except locale.Error:
        print("Aviso: Locales de português não encontrados. Usando o padrão do sistema.")


def buscar_aluno(df_alunos, matricula_pdf=None, nome_pdf=None, logger=print):
    """
    Busca um aluno no DataFrame df_alunos usando uma lógica de cascata aprimorada
    para lidar com nomes parcialmente cortados e evitar colisões.
    """
    # Nível 1: Busca por matrícula (mais confiável)
    if matricula_pdf:
        resultado = df_alunos[df_alunos['matricula'] == str(matricula_pdf)]
        if not resultado.empty:
            return resultado.iloc[0]

    if nome_pdf:
        # Nível 2: Busca por nome completo exato
        nome_pdf_normalizado = ' '.join(nome_pdf.strip().upper().split())
        resultado = df_alunos[df_alunos['nome'].str.strip().str.upper() == nome_pdf_normalizado]
        if not resultado.empty:
            return resultado.iloc[0]

        # --- NÍVEL 3: BUSCA PARCIAL COM VERIFICAÇÃO DE AMBIGUIDADE ---
        palavras_pdf = nome_pdf_normalizado.split()
        
        if len(palavras_pdf) < 2:
            return None

        correspondencias_parciais = []

        for index, row in df_alunos.iterrows():
            nome_db_normalizado = ' '.join(row['nome'].strip().upper().split())
            palavras_db = nome_db_normalizado.split()

            if len(palavras_pdf) > len(palavras_db):
                continue
            
            match = True
            for i in range(len(palavras_pdf)):
                palavra_pdf = palavras_pdf[i]
                palavra_db = palavras_db[i]

                if i == len(palavras_pdf) - 1:
                    if not palavra_db.startswith(palavra_pdf):
                        match = False
                        break
                else:
                    if palavra_pdf != palavra_db:
                        match = False
                        break
            
            if match:
                correspondencias_parciais.append(row)

        # --- Análise dos resultados da busca parcial ---
        
        if len(correspondencias_parciais) == 1:
            aluno_encontrado = correspondencias_parciais[0]
            logger(f"  -> Correspondência parcial única encontrada: '{nome_pdf}' -> '{aluno_encontrado['nome']}'")
            return aluno_encontrado
        
        elif len(correspondencias_parciais) > 1:
            logger(f"  AVISO: Múltiplos alunos encontrados para o nome '{nome_pdf}'. Verificação manual necessária.")
            nomes_encontrados = [aluno['nome'] for aluno in correspondencias_parciais]
            logger(f"     -> Possíveis correspondências: {nomes_encontrados}")
            return None
        
    return None


def carregar_dados_base(logger):
    logger("Conectando aos bancos de dados...")
    try:
        script_dir = os.path.dirname(__file__)
        project_root = os.path.dirname(script_dir)
        path_alunos_db = os.path.join(project_root, 'db', 'alunos.db')
        path_horarios_db = os.path.join(project_root, 'db', 'horarios.db')
        conn_alunos = sqlite3.connect(path_alunos_db)
        df_alunos = pd.read_sql_query("SELECT * FROM alunos", conn_alunos)
        conn_alunos.close()
        conn_horarios = sqlite3.connect(path_horarios_db)
        df_horarios = pd.read_sql_query("SELECT * FROM horarios", conn_horarios)
        conn_horarios.close()
        df_alunos['matricula'] = df_alunos['matricula'].astype(str)
        df_horarios['hora_inicio'] = pd.to_datetime(df_horarios['hora_inicio'], format='%H:%M').dt.time
        df_horarios['hora_fim'] = pd.to_datetime(df_horarios['hora_fim'], format='%H:%M').dt.time
        logger("Bancos de dados carregados com sucesso.")
        return df_alunos, df_horarios
    except Exception as e:
        logger(f"ERRO ao carregar os bancos de dados: {e}")
        return None, None


# Em modulos/logica.py, substitua apenas esta função:

def processar_dados_diarios(ausentes_path, frequencia_path, logger):
    from modulos.extrator_ausentes import extrair_dados_ausentes
    from modulos.extrator_frequencias import extrair_dados_frequencia

    df_alunos, df_horarios = carregar_dados_base(logger)
    if df_alunos is None or df_horarios is None:
        return None, None, None

    faltas_registradas = {}
    problemas_alunos = []

    logger("\n--- Processando Relatório de Ausentes ---")
    df_ausentes, report_date = extrair_dados_ausentes(ausentes_path)

    if report_date is None:
        logger("ERRO: Não foi possível encontrar a data no PDF de ausentes.")
        return None, None, None
    
    # --- INÍCIO DA CORREÇÃO DE LÓGICA DO DIA DA SEMANA ---
    
    # Pega o número do dia da semana (0=Segunda, 1=Terça, etc.)
    dia_numero = report_date.weekday()
    
    # Mapeamento direto por número do dia (mais confiável que strings com encoding)
    dias_semana_map = {
        0: 'SEGUNDA-FEIRA', 1: 'TERÇA-FEIRA', 2: 'QUARTA-FEIRA',
        3: 'QUINTA-FEIRA', 4: 'SEXTA-FEIRA', 5: 'SÁBADO', 6: 'DOMINGO'
    }
    
    dia_semana = dias_semana_map.get(dia_numero)
    if dia_semana is None:
        logger(f"ERRO: Dia da semana desconhecido (índice: {dia_numero}).")
        return None, None, None

    logger(f"Data do relatório identificada: {report_date.date()} ({dia_semana})")
    
    # --- FIM DA CORREÇÃO ---
    
    if df_ausentes is not None and not df_ausentes.empty:
        logger(f"Encontrados {len(df_ausentes)} alunos ausentes.")
        for _, row in df_ausentes.iterrows():
            matricula_pdf, nome_pdf = row['Matrícula'], row['Nome']
            info_aluno = buscar_aluno(df_alunos, matricula_pdf=matricula_pdf, nome_pdf=nome_pdf, logger=logger)
            if info_aluno is not None:
                turma, nome_db, matricula_db = info_aluno['turma'], info_aluno['nome'], info_aluno['matricula']
                problemas_alunos.append({'Matricula': matricula_db, 'Nome do Aluno': nome_db, 'Turma': turma, 'Problema': 'FALTOU', 'Acesso': 'Sem registro'})
                # A busca agora usa o 'dia_semana' corrigido
                aulas_do_dia = df_horarios[(df_horarios['turma'] == turma) & (df_horarios['dia_semana'] == dia_semana)]
                for _, aula in aulas_do_dia.iterrows():
                    chave_falta = (matricula_db, nome_db, turma, aula['disciplina'])
                    faltas_registradas[chave_falta] = faltas_registradas.get(chave_falta, 0) + 1
            else:
                logger(f"  Aviso: Aluno ausente '{nome_pdf.strip()}' não encontrado no BD.")

    logger("\n--- Processando Relatório de Frequência ---")
    df_frequencia, date_frequencia = extrair_dados_frequencia(frequencia_path)

    # Validação de datas
    if date_frequencia is None or report_date.date() != date_frequencia.date():
        logger("\n!! ERRO CRÍTICO !! As datas dos PDFs de ausentes e frequência não coincidem.")
        return None, None, None

    if df_frequencia is not None and not df_frequencia.empty:
        df_frequencia['Hora'] = pd.to_datetime(df_frequencia['Hora'], format='%H:%M:%S').dt.time
        for grupo_keys, acesso_aluno_df in df_frequencia.groupby(['Crachá', 'Nome']):
            matricula_pdf, nome_pdf = grupo_keys
            info_aluno = buscar_aluno(df_alunos, matricula_pdf=matricula_pdf, nome_pdf=nome_pdf, logger=logger)
            if info_aluno is None:
                logger(f"  Aviso: Aluno presente '{nome_pdf.strip()}' não encontrado no BD.")
                continue

            turma, nome_db, matricula_db = info_aluno['turma'], info_aluno['nome'], info_aluno['matricula']
            aulas_do_dia = df_horarios[(df_horarios['turma'] == turma) & (df_horarios['dia_semana'] == dia_semana)]
            
            if not aulas_do_dia.empty:
                primeira_aula_do_dia = aulas_do_dia['hora_inicio'].min()
                ultima_aula_do_dia = aulas_do_dia['hora_fim'].max()
                
                primeira_entrada = acesso_aluno_df[acesso_aluno_df['Sentido'] == 'Entrada']['Hora'].min()
                ultima_saida = acesso_aluno_df[acesso_aluno_df['Sentido'] == 'Saída']['Hora'].max()

                if pd.notna(primeira_entrada) and primeira_entrada > primeira_aula_do_dia:
                    problemas_alunos.append({'Matricula': matricula_db, 'Nome do Aluno': nome_db, 'Turma': turma, 'Problema': 'CHEGOU ATRASADO', 'Acesso': f"Entrada: {primeira_entrada.strftime('%H:%M:%S')}"})
                
                if pd.notna(ultima_saida) and ultima_saida < ultima_aula_do_dia:
                    problemas_alunos.append({'Matricula': matricula_db, 'Nome do Aluno': nome_db, 'Turma': turma, 'Problema': 'SAIU CEDO', 'Acesso': f"Saída: {ultima_saida.strftime('%H:%M:%S')}"})

                if pd.notna(primeira_entrada):
                    for _, aula in aulas_do_dia[aulas_do_dia['hora_fim'] < primeira_entrada].iterrows():
                        chave_falta = (matricula_db, nome_db, turma, aula['disciplina'])
                        faltas_registradas[chave_falta] = faltas_registradas.get(chave_falta, 0) + 1
                
                if pd.notna(ultima_saida):
                    for _, aula in aulas_do_dia[aulas_do_dia['hora_inicio'] > ultima_saida].iterrows():
                        chave_falta = (matricula_db, nome_db, turma, aula['disciplina'])
                        faltas_registradas[chave_falta] = faltas_registradas.get(chave_falta, 0) + 1
    
    df_problemas = pd.DataFrame(problemas_alunos)
    return report_date, faltas_registradas, df_problemas


def gerar_relatorio_faltas(dados_da_sessao, logger):
    """
    Lê o arquivo Excel existente, mescla com os novos dados da sessão atual,
    e gera o arquivo Excel completo, com todas as abas e o resumo atualizado.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook

    logger("\n--- Gerando Relatório Detalhado de Faltas com Resumo Semanal ---")
    if not dados_da_sessao:
        logger("Nenhum dado novo foi processado para gerar relatório.")
        return False
    
    logger(f"Dias processados na sessão atual: {list(dados_da_sessao.keys())}")
    for dia, (data, faltas, problemas) in dados_da_sessao.items():
        logger(f"  - {dia}: {len(faltas) if faltas else 0} registros de faltas")

    sheet_name_resumo = 'Quantitativo Total da Semana'
    script_dir = os.path.dirname(__file__)
    project_root = os.path.dirname(script_dir)
    relatorios_dir = os.path.join(project_root, 'relatorios')
    os.makedirs(relatorios_dir, exist_ok=True)
    output_path = os.path.join(relatorios_dir, 'relatorio_faltas_detalhado.xlsx')

    # 1. Carrega os dados de TODAS as abas do arquivo Excel existente.
    dados_existentes = {}
    if os.path.exists(output_path):
        try:
            dados_existentes = pd.read_excel(output_path, sheet_name=None)
            if sheet_name_resumo in dados_existentes:
                del dados_existentes[sheet_name_resumo]
            logger(f"Arquivo existente carregado. Abas encontradas: {list(dados_existentes.keys())}")
        except Exception as e:
            logger(f"Aviso: Não foi possível ler o arquivo Excel existente. Ele será sobrescrito. Erro: {e}")

    # 2. Converte os novos dados da sessão para DataFrames e mescla com dados existentes
    logger(f"\nProcessando {len(dados_da_sessao)} dia(s) da sessão atual...")
    for sheet_name, (report_date, faltas_dict, _) in dados_da_sessao.items():
        logger(f"Processando dia: {sheet_name}")
        if faltas_dict:
            lista_faltas = [list(chave) + [valor] for chave, valor in faltas_dict.items()]
            dados_existentes[sheet_name] = pd.DataFrame(lista_faltas, columns=['Matricula', 'Nome', 'Turma', 'Disciplina', 'Total de Faltas'])
            logger(f"  ✓ Dados do dia {sheet_name} adicionados/atualizados com {len(lista_faltas)} registros.")
        else:
            dados_existentes[sheet_name] = pd.DataFrame(columns=['Matricula', 'Nome', 'Turma', 'Disciplina', 'Total de Faltas'])
            logger(f"  ⚠ Dia {sheet_name} adicionado sem registros de faltas (possível erro no processamento).")

    logger(f"Total de abas a serem escritas: {len(dados_existentes)}")

    # 3. Calcula o resumo a partir do conjunto completo de dados
    df_resumo = None
    if dados_existentes:
        dfs_para_resumo = []
        for df in dados_existentes.values():
            if 'STATUS' in df.columns:
                dfs_para_resumo.append(df.drop(columns=['STATUS']))
            else:
                dfs_para_resumo.append(df)
        
        if dfs_para_resumo:
            combined_df = pd.concat(dfs_para_resumo, ignore_index=True)
            df_resumo = combined_df.groupby(['Matricula', 'Nome', 'Turma', 'Disciplina'])['Total de Faltas'].sum().reset_index()
            df_resumo.rename(columns={'Total de Faltas': 'Total na Semana'}, inplace=True)
            df_resumo.sort_values(by=['Turma', 'Nome', 'Disciplina'], inplace=True)
            df_resumo['STATUS'] = 'PENDENTE'

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
            for sheet_name in sorted(dados_existentes.keys()):
                df = dados_existentes[sheet_name].copy()
                df.sort_values(by=['Turma', 'Nome', 'Disciplina'], inplace=True)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                logger(f"Aba '{sheet_name}' escrita com {len(df)} registros.")
            
            if df_resumo is not None and not df_resumo.empty:
                df_resumo.to_excel(writer, sheet_name=sheet_name_resumo, index=False)
                logger(f"Aba de resumo '{sheet_name_resumo}' escrita com {len(df_resumo)} registros.")

        workbook = load_workbook(output_path)
        
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal='center')
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.coordinate in ws.merged_cells:
                        continue
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            if sheet_name == sheet_name_resumo:
                status_col_letter = get_column_letter(ws.max_column)
                dv = DataValidation(
                    type="list",
                    formula1='"PENDENTE,LANÇADO"',
                    allow_blank=False
                )
                dv.error = 'Valor inválido'
                dv.errorTitle = 'Entrada inválida'
                dv.prompt = 'Selecione: PENDENTE ou LANÇADO'
                dv.promptTitle = 'Status da Falta'
                
                ws.add_data_validation(dv)
                dv.add(f'{status_col_letter}2:{status_col_letter}{ws.max_row}')
                
                yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                for row in range(2, ws.max_row + 1):
                    status_cell = ws[f'{status_col_letter}{row}']
                    if status_cell.value == 'PENDENTE':
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = yellow_fill
                    elif status_cell.value == 'LANÇADO':
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = green_fill
            
            ws.freeze_panes = 'A2'
        
        workbook.save(output_path)
        logger(f"\nRelatório detalhado salvo com sucesso em: {output_path}")
        logger(f"Abas criadas: {', '.join(sorted(dados_existentes.keys()))}")
        if df_resumo is not None:
            logger(f"Aba de resumo: {sheet_name_resumo}")
        return True
        
    except Exception as e:
        logger(f"ERRO ao salvar e formatar o arquivo Excel: {e}")
        import traceback
        logger(f"Detalhes do erro:\n{traceback.format_exc()}")
        return False


# Em modulos/logica.py, substitua apenas esta função:

def gerar_relatorio_simples(df_problemas, report_date, logger):
    logger("\n--- Gerando Relatório Simples de Frequência ---")
    if df_problemas.empty:
        logger("Nenhum problema de frequência encontrado para gerar o relatório simples.")
        return

    script_dir = os.path.dirname(__file__)
    project_root = os.path.dirname(script_dir)
    relatorios_dir = os.path.join(project_root, 'relatorios')
    os.makedirs(relatorios_dir, exist_ok=True)
    output_path = os.path.join(relatorios_dir, f"relatorio_frequencia_{report_date.strftime('%d%m%y')}.xlsx")

    # --- INÍCIO DA CORREÇÃO DE ENCODING ---
    # Mapeamento manual para evitar problemas de encoding/locale, com a capitalização correta.
    dias_semana_pt = {
        0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira',
        3: 'Quinta-feira', 4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'
    }
    dia_numero = report_date.weekday()
    dia_da_semana_pt = dias_semana_pt.get(dia_numero, '') # Pega o nome correto
    # --- FIM DA CORREÇÃO ---

    # (O resto do código de formatação é o mesmo)
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    turma_font = Font(name='Calibri', size=12, bold=True)
    turma_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True)
    
    fill_faltou = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_atrasado = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_saiu_cedo = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = report_date.strftime('%d-%m-%Y')

    # A linha do título agora usa a variável corrigida
    ws['A1'] = f"Relatório de Frequência - {dia_da_semana_pt}, {report_date.strftime('%d/%m/%Y')}"
    ws.merge_cells('A1:D1')
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    current_row = 3
    
    df_problemas.sort_values(by=['Turma', 'Nome do Aluno'], inplace=True)
    for turma, data in df_problemas.groupby('Turma'):
        # (O resto do código de escrita e formatação de células permanece o mesmo)
        for col_num in range(1, 5):
            ws.cell(row=current_row, column=col_num).border = thin_border
        
        cell_turma = ws[f'A{current_row}']
        cell_turma.value = f"TURMA: {turma}"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        cell_turma.font = turma_font
        cell_turma.fill = turma_fill
        current_row += 1

        headers = ['Matrícula', 'Nome do Aluno', 'Problema', 'Acesso']
        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header_text)
            cell.font = header_font
            cell.border = thin_border
        current_row += 1

        for _, aluno in data.iterrows():
            for col_num in range(1, 5):
                cell = ws.cell(row=current_row, column=col_num)
                if col_num == 1: cell.value = aluno['Matricula']
                elif col_num == 2: cell.value = aluno['Nome do Aluno']
                elif col_num == 3: cell.value = aluno['Problema']
                elif col_num == 4: cell.value = aluno['Acesso']
                
                cell.border = thin_border
                
                if aluno['Problema'] == 'FALTOU': cell.fill = fill_faltou
                elif aluno['Problema'] == 'CHEGOU ATRASADO': cell.fill = fill_atrasado
                elif aluno['Problema'] == 'SAIU CEDO': cell.fill = fill_saiu_cedo
            
            current_row += 1
        
        current_row += 1

    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 4) if col_idx == 2 else (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save(output_path)
        logger(f"Relatório simples formatado salvo com sucesso em '{output_path}'")
    except Exception as e:
        logger(f"ERRO ao salvar relatório simples formatado: {e}")