from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Em modulos/formatador_excel.py, substitua esta função:

def formatar_relatorio_simples(worksheet, df_problemas): # Mude 'ws' para 'worksheet' para clareza
    """Aplica a formatação completa ao relatório simples de frequência."""
    
    ws = worksheet # Atribui o worksheet recebido à variável 'ws' para o resto do código funcionar

    # Define estilos
    # ... (o resto da função, que já estava correta, permanece o mesmo) ...
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    turma_font = Font(name='Calibri', size=12, bold=True)
    turma_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True)
    
    fill_faltou = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_atrasado = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_saiu_cedo = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Aplica formatação aos headers e linhas
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center')

    # Encontra as linhas que precisam de formatação de turma e header
    turma_rows = []
    header_rows = []
    for row_idx, cell in enumerate(ws['A'], 1):
        if cell.value and str(cell.value).startswith("TURMA:"):
            turma_rows.append(row_idx)
            header_rows.append(row_idx + 1)
            
    # Formata os headers de turma
    for row_idx in turma_rows:
        ws[f'A{row_idx}'].font = turma_font
        ws[f'A{row_idx}'].fill = turma_fill

    # Formata os headers de coluna
    for row_idx in header_rows:
        for cell in ws[row_idx]:
            cell.font = header_font

    # Aplica bordas e cores às células de dados
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        problema_cell_value = ws[f'C{row[0].row}'].value
        fill_color = None
        if problema_cell_value == 'FALTOU': fill_color = fill_faltou
        elif problema_cell_value == 'CHEGOU ATRASADO': fill_color = fill_atrasado
        elif problema_cell_value == 'SAIU CEDO': fill_color = fill_saiu_cedo

        for cell in row:
            if cell.value is not None:
                cell.border = thin_border
                if fill_color and cell.column <= 4:
                     cell.fill = fill_color

    # Ajuste de colunas
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.coordinate in ws.merged_cells: continue
            if cell.value:
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
        adjusted_width = (max_length + 4) if col_idx == 2 else (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def formatar_relatorio_detalhado(worksheet, has_status_col=False):
    """Aplica a formatação ao relatório detalhado e de resumo."""
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
    
    for col_idx, col in enumerate(worksheet.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    
    worksheet.auto_filter.ref = worksheet.dimensions

    if has_status_col:
        status_col_letter = get_column_letter(worksheet.max_column)
        dv = DataValidation(type="list", formula1='"PENDENTE,LANÇADA"', allow_blank=True)
        worksheet.add_data_validation(dv)
        validation_range = f'{status_col_letter}2:{status_col_letter}{worksheet.max_row}'
        dv.add(validation_range)
        
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        worksheet.conditional_formatting.add(validation_range, FormulaRule(formula=[f'LEFT({status_col_letter}2, 8)="PENDENTE"'], fill=red_fill))
        worksheet.conditional_formatting.add(validation_range, FormulaRule(formula=[f'LEFT({status_col_letter}2, 7)="LANÇADA"'], fill=green_fill))