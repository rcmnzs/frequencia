import pandas as pd
import os
from openpyxl import load_workbook, Workbook
import sys
from datetime import time

# Adiciona o diretório raiz ao path para encontrar o 'config'
project_root_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(project_root_path)

import config

# --- FUNÇÃO DE FORMATAÇÃO DO RELATÓRIO DETALHADO ---
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def formatar_relatorio_detalhado(worksheet, has_status_col=False):
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
    for col_idx, col in enumerate(worksheet.columns, 1):
        try:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        except (ValueError, TypeError):
            pass
    worksheet.auto_filter.ref = worksheet.dimensions
    if has_status_col:
        status_col_letter = get_column_letter(worksheet.max_column)
        dv = DataValidation(type="list", formula1='"PENDENTE,LANÇADA"', allow_blank=True)
        worksheet.add_data_validation(dv)
        validation_range = f'{status_col_letter}2:{status_col_letter}{worksheet.max_row}'
        dv.add(validation_range)
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        worksheet.conditional_formatting.add(validation_range, FormulaRule(formula=[f'LEFT({status_col_letter}2, 8)="PENDENTE"'], fill=red_fill))
        worksheet.conditional_formatting.add(validation_range, FormulaRule(formula=[f'LEFT({status_col_letter}2, 7)="LANÇADA"'], fill=green_fill))

# --- FUNÇÃO DE GERAÇÃO DO RELATÓRIO DETALHADO ---
def gerar_relatorio_faltas(dados_da_sessao, logger):
    logger("\n--- Gerando Relatório Detalhado de Faltas com Resumo Semanal ---")
    if not dados_da_sessao:
        logger("Nenhum dado novo foi processado para gerar relatório.")
        return
    sheet_name_resumo = 'Quantitativo Total da Semana'
    output_path = os.path.join(config.REPORTS_DIR, config.DETAILED_REPORT_FILENAME)
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    dados_existentes = {}
    if os.path.exists(output_path):
        try:
            dados_existentes = pd.read_excel(output_path, sheet_name=None)
            if sheet_name_resumo in dados_existentes:
                del dados_existentes[sheet_name_resumo]
        except Exception as e:
            logger(f"Aviso: Não foi possível ler arquivo existente. Erro: {e}")
    for sheet_name, (report_date, faltas_dict, _) in dados_da_sessao.items():
        if faltas_dict:
            lista_faltas = [list(chave) + [valor] for chave, valor in faltas_dict.items()]
            dados_existentes[sheet_name] = pd.DataFrame(lista_faltas, columns=['Matricula', 'Nome', 'Turma', 'Disciplina', 'Total de Faltas'])
    if dados_existentes:
        dfs_para_resumo = [df for df in dados_existentes.values() if 'Total de Faltas' in df.columns and not df.empty]
        if dfs_para_resumo:
            combined_df = pd.concat(dfs_para_resumo, ignore_index=True)
            df_resumo = combined_df.groupby(['Matricula', 'Nome', 'Turma', 'Disciplina'])['Total de Faltas'].sum().reset_index()
            df_resumo.rename(columns={'Total de Faltas': 'Total na Semana'}, inplace=True)
            df_resumo['STATUS'] = 'PENDENTE'
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name in sorted(dados_existentes.keys()):
                df = dados_existentes[sheet_name]
                df.sort_values(by=['Turma', 'Nome', 'Disciplina'], inplace=True)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            if 'df_resumo' in locals() and not df_resumo.empty:
                df_resumo.sort_values(by=['Turma', 'Nome', 'Disciplina'], inplace=True)
                df_resumo.to_excel(writer, sheet_name=sheet_name_resumo, index=False)
        workbook = load_workbook(output_path)
        for sheet_name in workbook.sheetnames:
            formatar_relatorio_detalhado(workbook[sheet_name], has_status_col=(sheet_name == sheet_name_resumo))
        workbook.save(output_path)
        logger(f"Relatório detalhado salvo/atualizado com sucesso!")
    except Exception as e:
        logger(f"ERRO ao salvar relatório detalhado: {e}")

def gerar_relatorio_simples(df_problemas, report_date, logger):
    logger("\n--- Gerando Relatório Simples de Frequência ---")
    
    # *** FILTRO ADICIONADO AQUI: Só mostra acessos após 7:50 ***
    horario_corte = time(7, 50)
    
    def filtrar_por_horario(row):
        # Se for FALTOU, mantém (não tem horário específico)
        if row['Problema'] == 'FALTOU':
            return True
        
        # Para CHEGOU ATRASADO e SAIU CEDO, verifica o horário
        acesso_str = row['Acesso']
        if acesso_str and acesso_str != 'Sem registro':
            try:
                # Extrai o horário do formato "Entrada: HH:MM:SS" ou "Saída: HH:MM:SS"
                horario_str = acesso_str.split(': ')[1]  # Pega "HH:MM:SS"
                hora, minuto, segundo = map(int, horario_str.split(':'))
                horario_acesso = time(hora, minuto, segundo)
                
                # Só mostra se o acesso foi após 7:50
                return horario_acesso >= horario_corte
            except:
                # Se houver erro ao processar, mantém o registro
                return True
        return True
    
    # Aplica o filtro
    df_problemas_filtrado = df_problemas[df_problemas.apply(filtrar_por_horario, axis=1)].copy()
    
    if df_problemas_filtrado.empty:
        logger("Nenhum problema de frequência após 7:50 detectado. Relatório vazio.")
    else:
        logger(f"Registros filtrados (após 7:50): {len(df_problemas_filtrado)} de {len(df_problemas)} totais")
    
    file_name = f"{config.SIMPLE_REPORT_PREFIX}{report_date.strftime('%d%m%y')}.xlsx"
    output_path = os.path.join(config.REPORTS_DIR, file_name)
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    
    dias_semana_pt = {0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira', 
                      3: 'Quinta-feira', 4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'}
    dia_da_semana_pt = dias_semana_pt.get(report_date.weekday(), '')

    wb = Workbook()
    ws = wb.active
    ws.title = report_date.strftime('%d-%m-%Y')
    
    # Estilos
    title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    turma_font = Font(name='Calibri', size=12, bold=True)
    turma_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True)
    fill_faltou = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_atrasado = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_saiu_cedo = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Título Principal
    titulo_cell = ws.cell(row=1, column=1, value=f"Relatório de Frequência - {dia_da_semana_pt}, {report_date.strftime('%d/%m/%Y')}")
    titulo_cell.font = title_font
    titulo_cell.fill = title_fill
    titulo_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')
    for col in range(1, 5):
        ws.cell(row=1, column=col).border = thin_border
        ws.cell(row=1, column=col).font = title_font
        ws.cell(row=1, column=col).fill = title_fill
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
    
    current_row = 3
    
    # *** EXTRAI TURMAS DINAMICAMENTE DOS DADOS FILTRADOS ***
    if df_problemas_filtrado.empty:
        # Ainda salva o relatório com apenas o título
        wb.save(output_path)
        logger(f"Relatório simples salvo com sucesso em: {output_path}")
        return
    
    # Obtém lista única de turmas presentes nos dados filtrados, ordenadas
    turmas_com_problemas = sorted(df_problemas_filtrado['Turma'].unique())
    logger(f"Turmas com problemas detectados (após 7:50): {', '.join(turmas_com_problemas)}")
    
    for turma in turmas_com_problemas:
        # Filtra dados da turma
        data = df_problemas_filtrado[df_problemas_filtrado['Turma'] == turma].copy()
        
        # Header da Turma
        turma_cell = ws.cell(row=current_row, column=1, value=f"TURMA: {turma}")
        turma_cell.font = turma_font
        turma_cell.fill = turma_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).font = turma_font
            ws.cell(row=current_row, column=col).fill = turma_fill
        current_row += 1
        
        # Adiciona cabeçalho das colunas
        headers = ['Matrícula', 'Nome do Aluno', 'Problema', 'Acesso']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
        current_row += 1
        
        # Ordena e adiciona dados dos alunos
        data = data.sort_values(by=['Nome do Aluno'])
        for _, aluno in data.iterrows():
            problema = aluno['Problema']
            
            # Determina a cor de fundo
            fill_color = None
            if problema == 'FALTOU':
                fill_color = fill_faltou
            elif problema == 'CHEGOU ATRASADO':
                fill_color = fill_atrasado
            elif problema == 'SAIU CEDO':
                fill_color = fill_saiu_cedo
            
            # Adiciona a linha
            row_data = [aluno['Matricula'], aluno['Nome do Aluno'], 
                       aluno['Problema'], aluno['Acesso']]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border
                if fill_color:
                    cell.fill = fill_color
            
            current_row += 1
        
        # Linha em branco entre turmas
        current_row += 1

    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15  # Matrícula
    ws.column_dimensions['B'].width = 45  # Nome do Aluno
    ws.column_dimensions['C'].width = 18  # Problema
    ws.column_dimensions['D'].width = 15  # Acesso
        
    try:
        wb.save(output_path)
        logger(f"Relatório simples salvo com sucesso em: {output_path}")
    except Exception as e:
        logger(f"ERRO ao salvar relatório simples: {e}")